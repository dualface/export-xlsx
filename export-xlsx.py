# -*- coding: utf-8 -*-
import glob
import json
import os
import sys
from enum import Enum

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

"""
从 Excel 文件导出 JSON 文件

COPYRIGHT 2021 ALL RESERVED. (C) liaoyulei, https://github.com/dualface

github repo: https://github.com/dualface/export_xlsx
"""


class HeaderType(Enum):
    """定义列头的类型"""
    # 正常列头
    NORMAL = 1
    # 定义字典开始
    DICT_OPEN = 2
    # 定义字典结束
    DICT_CLOSE = 3
    # 定义数组开始
    ARRAY_OPEN = 4
    # 定义数组结束
    ARRAY_CLOSE = 5


class Header:
    """封装数据表格的单个列头"""

    def __init__(self, column, name, column_type, optional=False, anonymous=False):
        # 所在列
        self.column = column
        # 字段名
        self.name = name
        # 列头类型
        self.type = column_type
        # 是否是可选列
        self.optional = optional
        # 是否是索引
        self.index_order = 0
        # 是否是匿名字段
        self.anonymous = anonymous


class DocumentSchema:
    """规格定义"""

    def __init__(self, configs):
        for key in ("output", "index", "header_row", "first_data_row"):
            if key not in configs:
                raise KeyError(f"Schema(): not found {key} in configs")

        # 输出文件名
        self.output = configs["output"]
        # 索引列表
        self.index_names = list(map(str.strip, configs["index"].split(",")))
        if len(self.index_names) < 1:
            raise KeyError("must have least one index")
        if len(self.index_names) > 2:
            raise KeyError("at most have two indexes")

        # 列头所在行
        self.header_row = int(configs["header_row"])
        # 列头类型所在行
        self.header_type_row = int(configs["header_type_row"])
        # 列头所在的列
        if "header_col" in configs:
            self.header_col = int(configs["header_col"])
        else:
            self.header_col = 1
        # 数据起始行
        self.first_data_row = int(configs["first_data_row"])

        # 所有列头 [TableHeader]
        self.headers = []
        # 所有的字典定义 dict_name => [TableHeader, TableHeader, ...]
        self.dicts = dict()
        # 所有的数组定义 array_name => [TableHeader, TableHeader, ...]
        self.arrays = dict()

        # 添加列头时用于标记最后一个字典列名
        self._last_dict_name = None
        # 添加列头时用于标记最后一个数组列名
        self._last_array_name = None

    def dumps(self):
        """输出配置信息"""
        print("Schema:")
        print(f"    output: {self.output}")
        print(f"    indexes: {self.index_names}")
        print(f"    header_row: {self.header_row}")
        print(f"    header_type_row: {self.header_type_row}")
        print(f"    header_col: {self.header_col}")
        print(f"    first_data_row: {self.first_data_row}")
        print("")

        indent = ""
        for header in self.headers:
            optional = ""
            if header.optional:
                optional = " OPTIONAL"
            if header.type == HeaderType.DICT_OPEN:
                print(
                    f"column [{header.column:>2}]: {header.name}{optional} DICT {{")
                indent = "    "
            elif header.type == HeaderType.DICT_CLOSE:
                print(f"column [{header.column:>2}]: }}")
                indent = ""
            elif header.type == HeaderType.ARRAY_OPEN:
                print(
                    f"column [{header.column:>2}]: {header.name}{optional} ARRAY [")
                indent = "    "
            elif header.type == HeaderType.ARRAY_CLOSE:
                print(f"column [{header.column:>2}]: ]")
                indent = ""
            else:
                print(
                    f"column [{header.column:>2}]: {indent}{header.name}{optional}")
        print("")

    def add_header(self, column, name):
        """添加列头"""
        name = name.strip().replace(" ", "")
        anonymous = name[0] == "#"
        if anonymous:
            name = name[1:]

        last_char = name[len(name) - 1]
        if anonymous and last_char != "[":
            raise TypeError(f"only array can be anonymous")

        header_type = HeaderType.NORMAL

        if last_char == "{" or last_char == "[":
            name = name[0:len(name) - 1]

        optional = name[len(name) - 1] == "?"
        if optional:
            name = name[0:len(name) - 1]

        if last_char == "{":
            header_type = HeaderType.DICT_OPEN
            self._last_dict_name = name
            self.dicts[name] = []
        elif last_char == "}":
            header_type = HeaderType.DICT_CLOSE
            name = self._last_dict_name
        elif last_char == "[":
            header_type = HeaderType.ARRAY_OPEN
            self._last_array_name = name
            self.arrays[name] = []
        elif last_char == "]":
            header_type = HeaderType.ARRAY_CLOSE
            name = self._last_array_name

        header = Header(column, name, header_type,
                        optional=optional, anonymous=anonymous)
        if self._last_dict_name is not None:
            self.dicts[self._last_dict_name].append(header)
        elif self._last_array_name is not None:
            self.arrays[self._last_array_name].append(header)

        if last_char == "}":
            self._last_dict_name = None
        if last_char == "]":
            self._last_array_name = None

        self.headers.append(header)

    def add_index(self, index_name):
        """添加索引"""
        index_order = 1
        for header in self.headers:
            if header.name == index_name:
                header.index_order = index_order
                index_order = index_order + 1


class SheetCursor:
    """封装读取操作的光标位置"""

    def __init__(self, column, row):
        self.column = column
        self.row = row


class ExcelSheet:
    """封装对 Excel 工作表的操作"""

    def __init__(self, sheet):
        self.sheet = sheet
        self.grid = self._fetch_cells()
        self.schema = DocumentSchema(self._fetch_configs())
        self._fetch_headers()

    def load_records(self):
        """载入行

        1. 从 first_data_row, header_col 开始，往右顺序读取字段值。
        2. 当遇到 DICT_OPEN 或者 ARRAY_OPEN 时，则开始读取 DICT 或 ARRAY 定义的区域。
        3. 读取区域完成后，从当前行继续往右读取字段值。
        4. 最后构造包含当前记录所有字段的字典。

        """
        records = []
        cursor = SheetCursor(1, self.schema.first_data_row)
        while cursor.row <= self.sheet.max_row:
            if self._val(self.schema.header_col, cursor.row) is None:
                cursor.row = cursor.row + 1
                continue
            record = self._load_record(cursor)
            records.append(record)
        return records

    def make_indexed_records(self, records):
        """根据索引构建索引后的分组记录集"""
        indexed_rows = dict()
        last_index_name = self.schema.index_names[len(
            self.schema.index_names) - 1]
        for row in records:
            index_value = row[last_index_name]
            indexed_rows[index_value] = row

        if len(self.schema.index_names) == 1:
            return indexed_rows

        primary_indexed_rows = dict()
        primary_index_name = self.schema.index_names[0]
        for row in records:
            index_value = row[primary_index_name]
            if index_value not in primary_indexed_rows:
                primary_indexed_rows[index_value] = dict()
            group = primary_indexed_rows[index_value]
            group_index_value = row[last_index_name]
            group[group_index_value] = row

        return primary_indexed_rows

    # private

    @staticmethod
    def _convert_val(val):
        """转换单元格的值"""
        if val is None:
            return None
        val = str(val).strip()
        val_lower = val.lower()
        if val_lower == "null":
            return None
        elif val_lower == "true":
            return True
        elif val_lower == "false":
            return False
        elif str.isnumeric(val):
            return int(val)
        try:
            return round(float(val), 4)
        finally:
            return val

    def _val_with_coordinate(self, column, row):
        """返回指定单元格的值及单元格的坐标，如果有必要则转换为数字"""
        coordinate = get_column_letter(column) + str(row)
        return self._convert_val(self.grid[row][column]), coordinate

    def _val(self, column, row):
        """返回指定单元格的值，如果有必要则转换为数字"""
        return self._convert_val(self.grid[row][column])

    def _load_record(self, cursor):
        """载入一条记录

        1. 从 data_row 行的第一列开始，往右顺序读取字段值。
        2. 当遇到 DICT_OPEN 或者 ARRAY_OPEN 时，则开始读取 DICT 或 ARRAY 定义的区域。
           2.1. 如果是 ARRAY，则区域可能包括多行，以 ARRAY_CLOSE 标记结束区域
        3. 读取区域完成后，从 data_row 行继续往右读取字段值。
        4. 返回包含当前记录所有字段的字典。
        5. 返回记录字典，以及下一行记录的开始行
        """
        record = dict()
        # 读取每一个字段对应的值
        cursor.column = 1
        max_move_row = 1
        for header in self.schema.headers:
            if header.column < cursor.column:
                continue

            name = header.name
            if header.type == HeaderType.NORMAL:
                val = self._val(header.column, cursor.row)
                if (not header.optional) or (val is not None):
                    record[name] = val
                cursor.column = cursor.column + 1
            elif header.type == HeaderType.DICT_OPEN:
                val = self._fetch_dict(
                    self.schema.dicts[name], cursor, header.optional)
                if (not header.optional) or len(val) > 0:
                    record[name] = val
            elif header.type == HeaderType.ARRAY_OPEN:
                arr, read_rows_count = self._fetch_array(
                    self.schema.arrays[name], cursor, header.optional)
                if (not header.optional) or len(arr) > 0:
                    record[name] = arr
                if read_rows_count > max_move_row:
                    max_move_row = read_rows_count

        cursor.row = cursor.row + max_move_row
        return record

    def _fetch_dict(self, headers, cursor, optional):
        """读取当前行内指定的字典"""
        len_of_headers = len(headers)
        cursor.column = headers[len_of_headers - 1].column + 1

        val, coordinate = self._val_with_coordinate(
            headers[0].column, cursor.row)
        if val != "{":
            if optional:
                return dict()
            raise TypeError(f"cell at <{coordinate}> is not dict begin")
        val, coordinate = self._val_with_coordinate(
            headers[-1].column, cursor.row)
        if val != "}":
            raise TypeError(f"cell at <{coordinate}> is not dict end")

        dict_at_row = dict()
        for i in range(1, len_of_headers - 1):
            header = headers[i]
            val = self._val(header.column, cursor.row)
            if val is not None:
                dict_at_row[header.name] = val

        return dict_at_row

    def _fetch_array(self, headers, cursor, optional):
        """从光标位置开始读取包含多个字典的数组"""
        len_of_headers = len(headers)
        cursor.column = headers[len_of_headers - 1].column + 1

        val, coordinate = self._val_with_coordinate(
            headers[0].column, cursor.row)
        if val != "{" and val != "[":
            if optional:
                return [], 1
            raise TypeError(f"cell at <{coordinate}> is not array begin")

        arr = []
        read_rows_count = 0
        data_row = cursor.row
        anonymous = headers[0].anonymous
        while data_row <= self.sheet.max_row:
            dict_at_row = dict()
            arr_at_row = []
            for i in range(1, len_of_headers - 1):
                header = headers[i]
                val = self._val(header.column, data_row)
                if val is None:
                    continue

                if anonymous:
                    arr_at_row.append(val)
                else:
                    dict_at_row[header.name] = val

            if anonymous and len(arr_at_row) > 0:
                arr.extend(arr_at_row)
            elif len(dict_at_row) > 0:
                arr.append(dict_at_row)

            read_rows_count = read_rows_count + 1

            val = self._val(headers[-1].column, data_row)
            data_row = data_row + 1
            if val == "}" or val == "]":
                # 数组已经结束
                break

        return arr, read_rows_count

    def _fetch_configs(self):
        """从工作表中读取导出配置"""
        val = self._val(1, 1)
        if val is None:
            raise TypeError("not found configs")

        # 导出配置分为多行
        configs = dict()
        for line in val.split("\n"):
            # 每一行一个配置项 config_name: config_value
            parts = list(map(str.strip, line.split(":")))
            if len(parts) != 2:
                raise SyntaxError(f"invalid config line '{line}'")
            key, val = parts
            if str.isnumeric(val):
                configs[key] = int(val)
            else:
                configs[key] = val
        return configs

    def _fetch_headers(self):
        """从工作表中读取列头信息"""
        for column in range(self.schema.header_col, self.sheet.max_column + 1):
            name = self._val(column, self.schema.header_row)
            if name is None:
                continue
            self.schema.add_header(column, name)
        for index_name in self.schema.index_names:
            self.schema.add_index(index_name)

    def _fetch_cells(self):
        """将工作表的所有单元格全部载入内存，方便后续快速查询"""
        grid = dict()
        for row_index, row in enumerate(self.sheet.rows):
            row_in_grid = dict()
            grid[row_index + 1] = row_in_grid
            for col_index, cell in enumerate(row):
                row_in_grid[col_index + 1] = cell.value
        return grid


def print_help():
    print("""
usage: python3 export_xlsx.py FILENAME

examples:

    python3 export_xlsx.py test.xlsx
    python3 export_xlsx.py *.xlsx

""")


def load_all_rows_in_workbook(filename, verbose):
    """打开工作薄，遍历所有工作表，载入数据

    1. 遍历每一个工作表，读取工作表 A1 单元格
    2. 如果 A1 单元格不为空，则假定为工作表的导出设置
    3. 读取工作表内定义的列头
    4. 读取工作表的数据
    5. 每个工作表读取的数据会以输出文件名为 KEY 放入 all 字典
    6. 如果多个工作表使用相同的输出文件名，则会合并数据
    7. 最后返回 all 字典
    """
    print(f"load file '{os.path.basename(filename)}'")
    wb = load_workbook(filename=filename, data_only=True, read_only=True)

    # 从工作薄中载入的所有数据
    # filename => rows_dict
    all_rows = dict()
    for sheet_name in wb.sheetnames:
        sheet_instance = wb[sheet_name]
        if sheet_instance.max_row < 1 or sheet_instance["A1"].value is None:
            print(f"not found configs in sheet {sheet_name}")
            continue

        sheet = ExcelSheet(sheet_instance)
        if verbose:
            sheet.schema.dumps()
        records = sheet.load_records()
        indexed = sheet.make_indexed_records(records)
        name = sheet.schema.output
        if name in all_rows:
            for key in indexed:
                all_rows[name][key] = indexed[key]
        else:
            all_rows[name] = indexed

    if len(all_rows) == 0:
        print("skipped.")
        print("")

    return all_rows


def export_all_to_json(all_rows):
    """导出所有数据为 JSON 文件"""
    for output in all_rows:
        with open(output, "w") as f:
            print(f"write file '{output}'")
            f.write(json.dumps(all_rows[output], indent=4))
            print("")


def export_file(filename, verbose):
    all_rows = load_all_rows_in_workbook(filename, verbose)
    export_all_to_json(all_rows)


def export_files(names, verbose):
    for filename in glob.glob(names):
        basename = os.path.basename(filename)
        if basename[0] == "~" or basename[0] == ".":
            continue
        export_file(filename, verbose)


def main():
    if len(sys.argv) < 2:
        print_help()
        sys.exit(1)
    names = sys.argv[1]
    if len(sys.argv) > 2 and sys.argv[2] == "-q":
        verbose = False
    else:
        verbose = True
    export_files(names, verbose)


if __name__ == "__main__":
    main()
