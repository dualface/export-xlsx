# -*- coding: utf-8 -*-
import glob
import json
import os
import sys
import traceback
from enum import Enum
from enum import auto as EnumAuto

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

"""
从 Excel 文件导出 JSON 文件

COPYRIGHT 2021 ALL RESERVED. (C) liaoyulei, https://github.com/dualface

github repo: https://github.com/dualface/export_xlsx
"""


class TableHeaderType(Enum):
    """定义列头的类型"""
    # 正常列头
    NORMAL = EnumAuto()
    # 定义字典开始
    DICT_OPEN = EnumAuto()
    # 定义字典结束
    DICT_CLOSE = EnumAuto()
    # 定义数组开始
    ARRAY_OPEN = EnumAuto()
    # 定义数组结束
    ARRAY_CLOSE = EnumAuto()


class TableHeader:
    """封装数据表格的单个列头"""

    def __init__(self, column, name, type):
        # 所在列
        self.column = column
        # 字段名
        self.name = name
        # 列头类型
        self.type = type
        # 是否是索引
        self.index_order = 0


class TableHeaders:
    """封装数据表格的列头"""

    def __init__(self):
        # 所有列头 [TableHeader]
        self.headers = []
        # 所有的字典定义 dict_name => [TableHeader, TableHeader, ...]
        self.dicts = dict()
        # 所有的数组定义 array_name => [TableHeader, TableHeader, ...]
        self.arrays = dict()

        # 添加列头时用于标记最后一个字典列名
        self._last_dict_name = None
        # 添加列头时用于标记最后一个数组列名
        self._last_array_name = None

    def add(self, column, name):
        """添加列头"""
        name = name.strip()
        last_char = name[len(name)-1]
        header_type = TableHeaderType.NORMAL

        if last_char == "{":
            header_type = TableHeaderType.DICT_OPEN
            name = name[0:len(name)-1]
            self._last_dict_name = name
            self.dicts[name] = []
        elif last_char == "}":
            header_type = TableHeaderType.DICT_CLOSE
            name = self._last_dict_name
        elif last_char == "[":
            header_type = TableHeaderType.ARRAY_OPEN
            name = name[0:len(name)-1]
            self._last_array_name = name
            self.arrays[name] = []
        elif last_char == "]":
            header_type = TableHeaderType.ARRAY_CLOSE
            name = self._last_array_name

        header = TableHeader(column, name, header_type)
        if self._last_dict_name is not None:
            self.dicts[self._last_dict_name].append(header)
        elif self._last_array_name is not None:
            self.arrays[self._last_array_name].append(header)

        if last_char == "}":
            self._last_dict_name = None
        if last_char == "]":
            self._last_array_name = None

        self.headers.append(header)

    def add_index(self, index_name):
        """添加索引"""
        index_order = 1
        for header in self.headers:
            if header.name == index_name:
                header.index_order = index_order
                index_order = index_order + 1

    def dumps(self):
        """输出所有列头的信息"""
        indent = ""
        for header in self.headers:
            if header.type == TableHeaderType.DICT_OPEN:
                print(f"column [{header.column:>2}]: {header.name} DICT {{")
                indent = "    "
            elif header.type == TableHeaderType.DICT_CLOSE:
                print(f"column [{header.column:>2}]: }}")
                indent = ""
            elif header.type == TableHeaderType.ARRAY_OPEN:
                print(f"column [{header.column:>2}]: {header.name} ARRAY [")
                indent = "    "
            elif header.type == TableHeaderType.ARRAY_CLOSE:
                print(f"column [{header.column:>2}]: ]")
                indent = ""
            else:
                print(f"column [{header.column:>2}]: {indent}{header.name}")
        print("")


class SheetCursor:
    """封装读取操作的光标位置"""

    def __init__(self, column, row):
        self.column = column
        self.row = row


class ExcelSheet:
    """封装对 Excel 工作表的操作"""

    def __init__(self, sheet):
        # Excel 工作表
        self.sheet = sheet
        # 输出文件名
        self.output_filename = ""
        # 索引名
        # [index1, ...]
        self.index_names = []
        # 列头所在行
        self.header_row = 0
        # 数据起始行
        self.first_data_row = 0
        # 所有列头
        self.headers = TableHeaders()

        # 载入导出配置和列头
        self._fetch_configs()
        self._fetch_headers()

    def dumps_configs(self):
        """输出配置信息"""
        print("output_filename:", self.output_filename)
        print("indexes:", self.index_names)
        print("header_row:", self.header_row)
        print("first_data_row:", self.first_data_row)
        print("")

    def load_records(self):
        """载入行

        1. 从 first_data_row 行的第一列开始，往右顺序读取字段值。
        2. 当遇到 DICT_OPEN 或者 ARRAY_OPEN 时，则开始读取 DICT 或 ARRAY 定义的区域。
        3. 读取区域完成后，从当前行继续往右读取字段值。
        4. 最后构造包含当前记录所有字段的字典。

        """
        records = []
        cursor = SheetCursor(1, self.first_data_row)
        while cursor.row <= self.sheet.max_row:
            record = self._load_record(cursor)
            records.append(record)
        return records

    def make_indexed_records(self, records):
        """根据索引构建索引后的分组记录集"""
        indexed_rows = dict()
        last_index_name = self.index_names[len(self.index_names) - 1]
        for row in records:
            index_value = row[last_index_name]
            indexed_rows[index_value] = row

        if len(self.index_names) == 1:
            return indexed_rows

        primary_indexed_rows = dict()
        primary_index_name = self.index_names[0]
        for row in records:
            index_value = row[primary_index_name]
            if index_value not in primary_indexed_rows:
                primary_indexed_rows[index_value] = dict()
            group = primary_indexed_rows[index_value]
            group_index_value = row[last_index_name]
            group[group_index_value] = row

        return primary_indexed_rows

    # private

    def _convert_val(self, val):
        """转换单元格的值"""
        if val is None:
            return None
        val = str(val).strip()
        if val.lower() == "null":
            return None

        if str.isnumeric(val):
            return int(val)

        try:
            return float(val)
        except:
            pass
        return val

    def _val_with_coordinate(self, column, row):
        """返回指定单元格的值及单元格的坐标，如果有必要则转换为数字"""
        cell = self.sheet.cell(column=column, row=row)
        coordinate = get_column_letter(column) + str(row)
        return self._convert_val(cell.value), coordinate

    def _val(self, column, row):
        """返回指定单元格的值，如果有必要则转换为数字"""
        cell = self.sheet.cell(column=column, row=row)
        return self._convert_val(cell.value)

    def _load_record(self, cursor):
        """载入一条记录

        1. 从 data_row 行的第一列开始，往右顺序读取字段值。
        2. 当遇到 DICT_OPEN 或者 ARRAY_OPEN 时，则开始读取 DICT 或 ARRAY 定义的区域。
           2.1. 如果是 ARRAY，则区域可能包括多行，以 ARRAY_CLOSE 标记结束区域
        3. 读取区域完成后，从 data_row 行继续往右读取字段值。
        4. 返回包含当前记录所有字段的字典。
        5. 返回记录字典，以及下一行记录的开始行
        """
        record = dict()
        # 读取每一个字段对应的值
        cursor.column = 1
        max_move_row = 1
        for header in self.headers.headers:
            if header.column < cursor.column:
                continue

            name = header.name
            if header.type == TableHeaderType.NORMAL:
                record[name] = self._val(header.column, cursor.row)
                cursor.column = cursor.column + 1
            elif header.type == TableHeaderType.DICT_OPEN:
                record[name] = self._fetch_dict(
                    self.headers.dicts[name], cursor)
            elif header.type == TableHeaderType.ARRAY_OPEN:
                arr = self._fetch_array(self.headers.arrays[name], cursor)
                record[name] = arr
                if len(arr) > max_move_row:
                    max_move_row = len(arr)

        cursor.row = cursor.row + max_move_row
        return record

    def _fetch_dict(self, headers, cursor):
        """读取当前行内指定的字典"""
        len_of_headers = len(headers)
        val, coordinate = self._val_with_coordinate(
            headers[0].column, cursor.row)
        if val != "{":
            raise TypeError(f"cell at <{coordinate}> is not dict begin")
        val, coordinate = self._val_with_coordinate(
            headers[-1].column, cursor.row)
        if val != "}":
            raise TypeError(f"cell at <{coordinate}> is not dict end")

        dict_at_row = dict()
        for i in range(1, len_of_headers - 1):
            header = headers[i]
            dict_at_row[header.name] = self._val(header.column, cursor.row)

        cursor.column = headers[len_of_headers - 1].column+1
        return dict_at_row

    def _fetch_array(self, headers, cursor):
        """从光标位置开始读取包含多个字典的数组"""
        len_of_headers = len(headers)
        val, coordinate = self._val_with_coordinate(
            headers[0].column, cursor.row)
        if val != "{" and val != "[":
            raise TypeError(f"cell at <{coordinate}> is not array begin")

        arr = []
        data_row = cursor.row
        while data_row <= self.sheet.max_row:
            d = dict()
            for i in range(1, len_of_headers - 1):
                header = headers[i]
                d[header.name] = self._val(header.column, data_row)
            arr.append(d)

            val = self._val(headers[-1].column, data_row)
            data_row = data_row + 1
            if val == "}" or val == "]":
                # 数组已经结束
                break

        cursor.column = headers[len_of_headers - 1].column+1
        return arr

    def _fetch_configs(self):
        """从工作表中读取导出配置"""
        val = self._val(1, 1)
        if val is None:
            raise TypeError("not found configs")

        # 导出配置分为多行
        configs = dict()
        for line in val.split("\n"):
            # 每一行一个配置项 config_name: config_value
            parts = list(map(str.strip, line.split(":")))
            if len(parts) != 2:
                raise SyntaxError(f"invalid config line '{line}'")
            key, val = parts
            if str.isnumeric(val):
                configs[key] = int(val)
            else:
                configs[key] = val

        if ("output" not in configs) or (type(configs["output"]) is not str) or (len(configs["output"]) == 0):
            raise KeyError("not found config key 'output'")
        if ("index" not in configs) or (type(configs["index"]) is not str) or (len(configs["index"]) == 0):
            raise KeyError("not found config key 'index'")
        if ("header_row" not in configs) or (type(configs["header_row"]) is not int) or (configs["header_row"] <= 0):
            raise KeyError("not found config key 'header_row'")
        if ("first_data_row" not in configs) or (type(configs["first_data_row"]) is not int) or (configs["first_data_row"] <= 0):
            raise KeyError("not found config key 'first_data_row'")

        self.output_filename = configs["output"]
        self.index_names = list(map(str.strip, configs["index"].split(",")))
        if len(self.index_names) < 1:
            raise KeyError(
                "invalid config key 'index', must have least one name")
        if len(self.index_names) > 2:
            raise KeyError("invalid config key 'index', most have two name")

        self.header_row = configs["header_row"]
        self.first_data_row = configs["first_data_row"]

    def _fetch_headers(self):
        """从工作表中读取列头信息"""
        for column in range(1, self.sheet.max_column + 1):
            name = self._val(column, self.header_row)
            if name == None:
                continue
            self.headers.add(column, name)
        for index_name in self.index_names:
            self.headers.add_index(index_name)


def help():
    print("""
usage: python3 export_xlsx.py FILENAME

examples:

    python3 export_xlsx.py test.xlsx
    python3 export_xlsx.py *.xlsx

""")


def load_all_rows_in_workbook(filename):
    """打开工作薄，遍历所有工作表，载入数据

    1. 遍历每一个工作表，读取工作表的 A1 单元格
    2. 如果 A1 单元格不为空，则假定为工作表的导出设置
    3. 读取工作表内定义的列头
    4. 读取工作表的数据
    5. 每个工作表读取的数据会以输出文件名为 KEY 放入 all 字典
    6. 如果多个工作表使用相同的输出文件名，则会合并数据
    7. 最后返回 all 字典
    """
    print(f"load file '{os.path.basename(filename)}'")
    wb = load_workbook(filename=filename, read_only=True)

    # 从工作薄中载入的所有数据
    # filename => rows_dict
    all = dict()
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet.max_row < 1 or sheet["A1"].value is None:
            continue

        configSheet = ExcelSheet(sheet)
        configSheet.dumps_configs()
        configSheet.headers.dumps()
        records = configSheet.load_records()
        indexed = configSheet.make_indexed_records(records)
        name = configSheet.output_filename
        if name in all:
            for key in indexed:
                all[name][key] = indexed[key]
        else:
            all[name] = indexed

    return all


def export_all_to_json(all):
    """导出所有数据为 JSON 文件"""
    for output_filename in all:
        with open(output_filename, "w") as f:
            print(f"write file '{output_filename}'")
            f.write(json.dumps(all[output_filename], indent=4))
            print("")


def main():
    if len(sys.argv) < 2:
        help()
        sys.exit(1)

    names = sys.argv[1]
    for filename in glob.glob(names):
        all = load_all_rows_in_workbook(filename)
        export_all_to_json(all)


if __name__ == "__main__":
    main()
